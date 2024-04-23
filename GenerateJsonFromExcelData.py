import openpyxl
import json
import pandas as pd

from openpyxl import load_workbook
from json import dumps

data_file = 'data.xlsx'
workBook = load_workbook(data_file)
workSheet = workBook['data']
all_rows = list(workSheet.rows)
jsonList = []
for i in range(1, 4):
    jsonRow = {}
    for j in range(1, 3):
        columnName = workSheet.cell(row=1, column=j)
        rowData = workSheet.cell(row=i+1, column=j)
        jsonRow.update(
            {
                columnName.value : rowData.value
            }
        )
    jsonList.append(jsonRow)

json_data = dumps(jsonList)

with open('data.json', 'w') as jsonFile:
    json.dump(json_data, jsonFile)
